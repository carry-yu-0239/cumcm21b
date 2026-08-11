"""生成问题四的既定实验设计图表。

本程序只将既定的两阶段追加实验设计转化为论文图表：原始记录保持只读，
不补造四个新增实验或第 5 次复验的结果。运行：
    python src/q4/generate_q4_artifacts.py --strict
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path

import matplotlib
import numpy as np
import pandas as pd
from matplotlib.font_manager import FontProperties
from openpyxl import load_workbook

matplotlib.use("Agg")
import matplotlib.pyplot as plt


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
from src.q2.reproduce_q2 import load_data  # noqa: E402


RAW_PATH = ROOT / "data" / "raw" / "q1_attachment1.xlsx"
OUTPUT_DIR = ROOT / "outputs" / "q4"
TABLE_DIR = OUTPUT_DIR / "tables"
LOG_DIR = OUTPUT_DIR / "logs"
FIGURE_DIR = ROOT / "paper" / "figures"

EXPECTED = {
    "all": (("A3", 400.0, 44.720910), ("A3", 450.0, 43.113600), 1.607310),
    "low": (("A2", 325.0, 17.263556), ("A3", 325.0, 10.798720), 6.464836),
}


def displayed_conversion_values(path: Path) -> dict[tuple[str, float], float]:
    """按附件单元格的显示精度读取乙醇转化率。"""
    sheet = load_workbook(path, data_only=True, read_only=True).active
    values: dict[tuple[str, float], float] = {}
    group: str | None = None
    for row in sheet.iter_rows(min_row=2):
        if row[0].value is not None:
            group = str(row[0].value).strip()
        if group is None or row[2].value is None or row[3].value is None:
            continue
        number_format = str(row[3].number_format).split(";", 1)[0]
        decimals = len(number_format.split(".", 1)[1].split("_", 1)[0]) if "." in number_format else 0
        values[(group, float(row[2].value))] = round(float(row[3].value), decimals)
    if len(values) != 114:
        raise ValueError(f"附件 1 显示精度读取异常：获得 {len(values)} 条记录")
    return values


def latex_escape(value: object) -> str:
    return str(value).replace("&", r"\&").replace("%", r"\%").replace("_", r"\_")


def write_table(stem: str, headers: list[str], rows: list[list[object]], caption: str, label: str, columns: str) -> None:
    lines = [r"\begin{table}[H]", r"  \centering", rf"  \caption{{{caption}}}", rf"  \label{{{label}}}", r"  \small", rf"  \begin{{tabularx}}{{\textwidth}}{{{columns}}}", r"    \toprule"]
    lines.append("    " + " & ".join(headers) + r" \\")
    lines.append(r"    \midrule")
    lines.extend("    " + " & ".join(latex_escape(item) for item in row) + r" \\" for row in rows)
    lines += [r"    \bottomrule", r"  \end{tabularx}", r"\end{table}", ""]
    (TABLE_DIR / f"{stem}.tex").write_text("\n".join(lines), encoding="utf-8")


def plot_style() -> tuple[FontProperties, FontProperties]:
    """英文字体为 Times New Roman；中文使用宋体补足 Times 的缺失字形。"""
    font_dir = Path("C:/Windows/Fonts")
    times = FontProperties(fname=str(font_dir / "times.ttf"))
    chinese = FontProperties(fname=str(font_dir / "simsun.ttc"))
    plt.rcParams.update({
        "font.family": ["Times New Roman", "SimSun"],
        "axes.unicode_minus": False,
        "figure.dpi": 160,
        "savefig.dpi": 300,
    })
    return times, chinese


def chart_labels(axis: plt.Axes, title: str, chinese: FontProperties) -> None:
    axis.set_xlabel("温度 $T$ / ℃", fontproperties=chinese)
    axis.set_ylabel(r"$\mathrm{C}_4$ 烯烃收率 $Y_{\mathrm{C4}}$ / %", fontproperties=chinese)
    axis.set_title(title, fontproperties=chinese, pad=9)
    axis.grid(alpha=0.24, linewidth=0.7)


def plot_a3_design(records: pd.DataFrame, chinese: FontProperties) -> None:
    frame = records.loc[(records["group"] == "A3") & records["temperature_c"].isin([350, 400, 450])].sort_values("temperature_c")
    figure, axis = plt.subplots(figsize=(7.9, 4.9))
    axis.plot(frame["temperature_c"], frame["Y_C4_pct"], color="#A44136", marker="o", linewidth=2.0, label="A3 已有实测点")
    axis.axvline(425, color="#476F88", linestyle="--", linewidth=1.5)
    axis.annotate("E1：425 ℃\n单点二分加密\n收率待实测", xy=(425, 46.6), xytext=(426.5, 36.8), fontproperties=chinese, arrowprops={"arrowstyle": "->", "color": "#476F88"})
    axis.set_xlim(342, 458)
    axis.set_ylim(18, 48)
    chart_labels(axis, "A3 高收益温区的追加测点", chinese)
    axis.legend(prop=chinese, frameon=False, loc="lower left")
    figure.tight_layout()
    figure.savefig(FIGURE_DIR / "q4_a3_high_temperature_design.png", bbox_inches="tight")
    plt.close(figure)


def plot_a2_design(records: pd.DataFrame, chinese: FontProperties) -> None:
    frame = records.loc[(records["group"] == "A2") & records["temperature_c"].isin([300, 325, 350])].sort_values("temperature_c")
    figure, axis = plt.subplots(figsize=(7.9, 4.9))
    axis.plot(frame["temperature_c"], frame["Y_C4_pct"], color="#25776C", marker="o", linewidth=2.0, label="A2 已有实测点")
    axis.axvline(337.5, color="#476F88", linestyle="--", linewidth=1.5)
    axis.axvline(350, color="#7B2D26", linestyle=":", linewidth=1.5)
    axis.annotate("E2：337.5 ℃\n单点二分加密\n收率待实测", xy=(337.5, 29.7), xytext=(302, 22.5), fontproperties=chinese, arrowprops={"arrowstyle": "->", "color": "#476F88"})
    axis.annotate("严格约束 $T<350$ ℃", xy=(350, 26.54), xytext=(309, 28.3), fontproperties=chinese, arrowprops={"arrowstyle": "->", "color": "#7B2D26"})
    axis.set_xlim(296, 354)
    axis.set_ylim(5, 31)
    chart_labels(axis, "A2 严格低温区的追加测点", chinese)
    axis.legend(prop=chinese, frameon=False, loc="lower right")
    figure.tight_layout()
    figure.subplots_adjust(top=0.88)
    figure.savefig(FIGURE_DIR / "q4_a2_low_temperature_design.png", bbox_inches="tight")
    plt.close(figure)


def plot_co_profile(records: pd.DataFrame, chinese: FontProperties) -> None:
    mapping = {"A4": 0.5, "A1": 1.0, "A2": 2.0, "A6": 5.0}
    points = []
    for group, loading in mapping.items():
        for temperature in [350.0, 400.0]:
            match = records.loc[(records["group"] == group) & (records["temperature_c"] == temperature), "Y_C4_pct"]
            points.append((group, loading, temperature, None if match.empty else float(match.iloc[0])))
    figure, axis = plt.subplots(figsize=(7.9, 4.9))
    colors = {350.0: "#25776C", 400.0: "#A44136"}
    for temperature in [350.0, 400.0]:
        present = [(loading, value) for _, loading, temp, value in points if temp == temperature and value is not None]
        missing = [(loading, group) for group, loading, temp, value in points if temp == temperature and value is None]
        if temperature == 350.0:
            axis.plot([item[0] for item in present], [item[1] for item in present], color=colors[temperature], marker="o", linewidth=1.8, label=f"{temperature:.0f} ℃已有实测")
        else:
            axis.scatter([item[0] for item in present], [item[1] for item in present], color=colors[temperature], s=75, label=f"{temperature:.0f} ℃已有实测")
        for loading, group in missing:
            axis.scatter([loading], [0.035], transform=axis.get_xaxis_transform(), facecolors="white", edgecolors=colors[temperature], marker="s", s=55, zorder=4, clip_on=False)
            axis.annotate(f"{group} 未实验\n(E{'4' if group == 'A1' else '3'})", xy=(loading, 0.035), xycoords=axis.get_xaxis_transform(), xytext=(loading + 0.12, 5.4), textcoords="data", fontproperties=chinese, color=colors[temperature])
    axis.set_xticks([0.5, 1, 2, 5])
    axis.set_ylim(0, 42)
    axis.set_xlabel(r"Co 负载量 $\omega_{\mathrm{Co}}$ / wt%", fontproperties=chinese)
    axis.set_ylabel(r"$\mathrm{C}_4$ 烯烃收率 $Y_{\mathrm{C4}}$ / %", fontproperties=chinese)
    axis.set_title("Co 负载量的 350 ℃与 400 ℃响应结构", fontproperties=chinese, pad=9)
    axis.grid(alpha=0.24, linewidth=0.7)
    axis.legend(prop=chinese, frameon=False, loc="upper left")
    figure.tight_layout()
    figure.savefig(FIGURE_DIR / "q4_co_loading_profile_design.png", bbox_inches="tight")
    plt.close(figure)


def prepare_records() -> pd.DataFrame:
    records, _ = load_data(RAW_PATH)
    display_x = displayed_conversion_values(RAW_PATH)
    records["X_EtOH_display_pct"] = [display_x[(group, float(temperature))] for group, temperature in zip(records["group"], records["temperature_c"])]
    records["Y_C4_pct"] = records["X_EtOH_display_pct"] * records["S_C4_pct"] / 100.0
    return records


def rank(records: pd.DataFrame, scope: str) -> pd.DataFrame:
    frame = records if scope == "all" else records.loc[records["temperature_c"] < 350]
    return frame.sort_values(["Y_C4_pct", "group", "temperature_c"], ascending=[False, True, True]).reset_index(drop=True)


def make_tables(records: pd.DataFrame) -> dict[str, float]:
    all_rank, low_rank = rank(records, "all"), rank(records, "low")
    all_margin = float(all_rank.iloc[0]["Y_C4_pct"] - all_rank.iloc[1]["Y_C4_pct"])
    low_margin = float(low_rank.iloc[0]["Y_C4_pct"] - low_rank.iloc[1]["Y_C4_pct"])
    write_table("q4_initial_rankings", ["优化任务", "第一名", "$Y_{(1)}$/\\%", "第二名", "$Y_{(2)}$/\\%", "决策裕度/百分点"], [
        ["全温区", f"{all_rank.iloc[0]['group']}--{all_rank.iloc[0]['temperature_c']:.0f} ℃", f"{all_rank.iloc[0]['Y_C4_pct']:.6f}", f"{all_rank.iloc[1]['group']}--{all_rank.iloc[1]['temperature_c']:.0f} ℃", f"{all_rank.iloc[1]['Y_C4_pct']:.6f}", f"{all_margin:.6f}"],
        ["严格 $T<350$ ℃", f"{low_rank.iloc[0]['group']}--{low_rank.iloc[0]['temperature_c']:.0f} ℃", f"{low_rank.iloc[0]['Y_C4_pct']:.6f}", f"{low_rank.iloc[1]['group']}--{low_rank.iloc[1]['temperature_c']:.0f} ℃", f"{low_rank.iloc[1]['Y_C4_pct']:.6f}", f"{low_margin:.6f}"],
    ], "两类优化任务的当前观测排名与决策裕度", "tab:q4-initial-rankings", "L{0.17\\textwidth}C{0.16\\textwidth}C{0.13\\textwidth}C{0.16\\textwidth}C{0.13\\textwidth}C{0.15\\textwidth}")
    write_table("q4_phase_one_plan", ["实验", "条件", "温度", "类型与作用"], [
        ["E1", "A3 原条件", "425 ℃", "组内单点二分加密，细化 400--450 ℃高收益区"],
        ["E2", "A2 原条件", "337.5 ℃", "组内单点二分加密，细化严格低温区 325--350 ℃"],
        ["E3", "A2 原条件", "400 ℃", "受控高温扩展；补 2 wt% Co 高温缺格及 A2/A5 进料对照"],
        ["E4", "A1 原条件", "400 ℃", "受控高温扩展；补 1 wt% Co 高温缺格及 A1/A3 进料对照"],
    ], "第一阶段四次定向追加实验", "tab:q4-phase-one", "C{0.09\\textwidth}C{0.18\\textwidth}C{0.13\\textwidth}X")
    profiles = []
    for temperature in [350.0, 400.0]:
        row = [f"{temperature:.0f} ℃"]
        for group in ["A4", "A1", "A2", "A6"]:
            match = records.loc[(records["group"] == group) & (records["temperature_c"] == temperature), "Y_C4_pct"]
            row.append("未实验" if match.empty else f"{float(match.iloc[0]):.5f}")
        profiles.append(row)
    write_table("q4_co_loading_profile", ["温度", "0.5 wt\\% (A4)", "1 wt\\% (A1)", "2 wt\\% (A2)", "5 wt\\% (A6)"], profiles, "Co 负载量严格匹配块的现有收率结构", "tab:q4-co-profile", "C{0.14\\textwidth}C{0.18\\textwidth}C{0.18\\textwidth}C{0.18\\textwidth}C{0.18\\textwidth}")
    pair_rows = []
    for left, right, planned in [("A1", "A3", "E4：A1--400 ℃"), ("A2", "A5", "E3：A2--400 ℃")]:
        values = []
        for group in [left, right]:
            match = records.loc[(records["group"] == group) & (records["temperature_c"] == 400), "Y_C4_pct"]
            values.append("未实验" if match.empty else f"{float(match.iloc[0]):.5f}")
        pair_rows.append([f"{left}/{right}", "乙醇进料条件", values[0], values[1], planned])
    write_table("q4_feed_match_pairs", ["严格匹配对", "仅改变因素", "前者 400 ℃收率/\\%", "后者 400 ℃收率/\\%", "补全安排"], pair_rows, "400 ℃乙醇进料严格匹配对的补全安排", "tab:q4-feed-pairs", "C{0.15\\textwidth}C{0.17\\textwidth}C{0.18\\textwidth}C{0.18\\textwidth}C{0.18\\textwidth}")
    return {"all_margin": all_margin, "low_margin": low_margin}


def validate(records: pd.DataFrame, margins: dict[str, float]) -> dict[str, object]:
    all_rank, low_rank = rank(records, "all"), rank(records, "low")
    checks = {
        "record_count": len(records) == 114,
        "group_count": records["group"].nunique() == 21,
        "all_yield_in_range": bool(records["Y_C4_pct"].between(0, 100).all()),
        "all_ranking": (all_rank.iloc[0]["group"], float(all_rank.iloc[0]["temperature_c"]), round(float(all_rank.iloc[0]["Y_C4_pct"]), 6)) == EXPECTED["all"][0] and (all_rank.iloc[1]["group"], float(all_rank.iloc[1]["temperature_c"]), round(float(all_rank.iloc[1]["Y_C4_pct"]), 6)) == EXPECTED["all"][1],
        "low_ranking": (low_rank.iloc[0]["group"], float(low_rank.iloc[0]["temperature_c"]), round(float(low_rank.iloc[0]["Y_C4_pct"]), 6)) == EXPECTED["low"][0] and (low_rank.iloc[1]["group"], float(low_rank.iloc[1]["temperature_c"]), round(float(low_rank.iloc[1]["Y_C4_pct"]), 6)) == EXPECTED["low"][1],
        "all_margin": abs(margins["all_margin"] - EXPECTED["all"][2]) < 1e-6,
        "low_margin": abs(margins["low_margin"] - EXPECTED["low"][2]) < 1e-6,
        "400c_missing_only_a1_a2": set(records.loc[records["temperature_c"] == 400, "group"]) == set(records["group"].unique()) - {"A1", "A2"},
    }
    return {"status": "PASS" if all(checks.values()) else "DISCREPANCY", "checks": checks, "source_sha256": hashlib.sha256(RAW_PATH.read_bytes()).hexdigest(), "margins": margins}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--strict", action="store_true")
    args = parser.parse_args()
    for directory in [TABLE_DIR, LOG_DIR, FIGURE_DIR]:
        directory.mkdir(parents=True, exist_ok=True)
    records = prepare_records()
    margins = make_tables(records)
    _, chinese = plot_style()
    plot_a3_design(records, chinese)
    plot_a2_design(records, chinese)
    plot_co_profile(records, chinese)
    report = validate(records, margins)
    (LOG_DIR / "q4_validation.json").write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["status"] == "PASS" or not args.strict else 1


if __name__ == "__main__":
    raise SystemExit(main())
