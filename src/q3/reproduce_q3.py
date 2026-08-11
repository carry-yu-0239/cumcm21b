"""复现问题三既定的观测排序、局部连续化和资格检验。

本程序仅落实《问题3_模型交接单》给出的收率定义、候选代理资格检验、
局部分段线性规则与敏感性检查；不搜索新的模型、配方、参数或结论。

运行：python src/q3/reproduce_q3.py --strict
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import sys
from pathlib import Path
from typing import Iterable

import matplotlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
from src.q2.reproduce_q2 import COMBINATIONS, load_data  # noqa: E402


RAW_PATH = ROOT / "data" / "raw" / "q1_attachment1.xlsx"
HANDOFF_PATH = ROOT / "docs" / "handoffs" / "q3_model_handoff.docx"
OUTPUT_DIR = ROOT / "outputs" / "q3"
TABLE_DIR = OUTPUT_DIR / "tables"
FIGURE_DIR = OUTPUT_DIR / "figures"
LOG_DIR = OUTPUT_DIR / "logs"

EXPECTED = {
    "observed_best": ("A3", 400.0, 44.720910),
    "strict_observed_best": ("A2", 325.0, 17.263556),
    "strict_supremum": 26.541080,
    "quadratic_vertex_temperature": 422.16,
    "quadratic_vertex_yield": 47.50,
    "proxy_rows": 108,
    "proxy_groups": 20,
    "loco": {
        "temperature_only": (3, 5.9528, 3.5336, 17.99),
        "main_effects": (8, 5.2454, 3.7413, 24.25),
        "two_interactions": (10, 5.3886, 3.7351, 26.80),
    },
}


def group_order(series: pd.Series) -> pd.Series:
    return series.map({group: position for position, group in enumerate(COMBINATIONS)})


def displayed_conversion_values(path: Path) -> dict[tuple[str, float], float]:
    """Recover the workbook's displayed precision for the conversion column."""
    sheet = load_workbook(path, data_only=True, read_only=True).active
    values: dict[tuple[str, float], float] = {}
    group: str | None = None
    for row in sheet.iter_rows(min_row=2):
        if row[0].value is not None:
            group = str(row[0].value).strip()
        if group is None or row[2].value is None or row[3].value is None:
            continue
        number_format = str(row[3].number_format).split(";")[0]
        decimals = len(number_format.split(".", 1)[1].split("_", 1)[0]) if "." in number_format else None
        value = float(row[3].value)
        if decimals is not None:
            value = round(value, decimals)
        values[(group, float(row[2].value))] = value
    if len(values) != 114:
        raise ValueError(f"附件1显示精度读取异常：获得 {len(values)} 条转换率记录")
    return values


def latex_escape(value: object) -> str:
    return str(value).replace("&", r"\&").replace("%", r"\%").replace("_", r"\_")


def write_longtable(stem: str, headers: list[str], rows: Iterable[Iterable[object]], caption: str, label: str, column_spec: str) -> None:
    lines = [r"\begin{longtable}{" + column_spec + "}", rf"\caption{{{caption}}}\label{{{label}}}\\", r"\toprule"]
    lines.append(" & ".join(headers) + r" \\")
    lines += [r"\midrule", r"\endfirsthead", r"\toprule", " & ".join(headers) + r" \\", r"\midrule", r"\endhead"]
    lines.extend(" & ".join(latex_escape(value) for value in row) + r" \\" for row in rows)
    lines += [r"\bottomrule", r"\end{longtable}", ""]
    (TABLE_DIR / f"{stem}.tex").write_text("\n".join(lines), encoding="utf-8")


def style_plots() -> tuple[FontProperties, FontProperties]:
    """Use Times New Roman by default and SimSun only as a Chinese glyph fallback."""
    font_dir = Path("C:/Windows/Fonts")
    times = FontProperties(fname=str(font_dir / "times.ttf"))
    chinese = FontProperties(fname=str(font_dir / "simsun.ttc"))
    plt.rcParams.update({
        "font.family": ["Times New Roman", "SimSun"],
        "axes.unicode_minus": False,
        "figure.dpi": 160,
        "savefig.dpi": 300,
    })
    return times, chinese


def label(axis: plt.Axes, xlabel: str, ylabel: str, title: str, chinese: FontProperties) -> None:
    axis.set_xlabel(xlabel, fontproperties=chinese)
    axis.set_ylabel(ylabel, fontproperties=chinese)
    axis.set_title(title, fontproperties=chinese, pad=10)
    axis.grid(alpha=0.20, linewidth=0.7)


def plot_all_trajectories(records: pd.DataFrame, chinese: FontProperties) -> None:
    figure, axis = plt.subplots(figsize=(8.8, 5.35))
    for group in COMBINATIONS:
        frame = records.loc[records["group"] == group].sort_values("temperature_c")
        if group in {"A2", "A3"}:
            continue
        axis.plot(frame["temperature_c"], frame["Y_C4_pct"], color="#AEB8BC", linewidth=0.85, marker="o", markersize=2.4, alpha=0.72)
    for group, color, name in [("A2", "#267D72", "A2"), ("A3", "#B04436", "A3")]:
        frame = records.loc[records["group"] == group].sort_values("temperature_c")
        axis.plot(frame["temperature_c"], frame["Y_C4_pct"], color=color, linewidth=2.15, marker="o", markersize=4.2, label=name, zorder=3)
    label(axis, "温度 $T$ / ℃", r"$\mathrm{C}_4$ 烯烃收率 $Y_{\mathrm{C4}}$ / %", "各催化剂组合的实测收率--温度轨迹", chinese)
    axis.legend(prop=chinese, frameon=False, loc="upper left")
    figure.tight_layout()
    figure.savefig(FIGURE_DIR / "q3_all_yield_temperature.png", bbox_inches="tight")
    plt.close(figure)


def plot_a3_sensitivity(records: pd.DataFrame, chinese: FontProperties) -> tuple[float, float]:
    frame = records.loc[(records["group"] == "A3") & (records["temperature_c"].isin([350, 400, 450]))].sort_values("temperature_c")
    x = frame["temperature_c"].to_numpy(dtype=float)
    y = frame["Y_C4_pct"].to_numpy(dtype=float)
    coefficients = np.polyfit(x - 400.0, y, 2)
    vertex_offset = -coefficients[1] / (2 * coefficients[0])
    vertex_t = float(400.0 + vertex_offset)
    vertex_y = float(np.polyval(coefficients, vertex_offset))
    grid = np.linspace(350, 450, 301)

    figure, axis = plt.subplots(figsize=(8.1, 5.0))
    axis.plot(x, y, color="#B04436", marker="o", linewidth=2.0, label="A3 实测节点与低自由度连接")
    axis.plot(grid, np.polyval(coefficients, grid - 400.0), color="#495057", linewidth=1.45, linestyle="--", label="三点二次压力测试")
    axis.scatter([vertex_t], [vertex_y], color="#495057", s=28, zorder=4)
    axis.annotate(f"压力测试顶点\n({vertex_t:.2f} ℃, {vertex_y:.2f}%)", xy=(vertex_t, vertex_y), xytext=(356, 46.2), arrowprops={"arrowstyle": "->", "color": "#495057"}, fontproperties=chinese)
    label(axis, "温度 $T$ / ℃", r"$\mathrm{C}_4$ 烯烃收率 $Y_{\mathrm{C4}}$ / %", "A3 高温区局部模型形式敏感性", chinese)
    axis.legend(prop=chinese, frameon=False, loc="lower left")
    figure.tight_layout()
    figure.savefig(FIGURE_DIR / "q3_a3_high_temperature_sensitivity.png", bbox_inches="tight")
    plt.close(figure)
    return vertex_t, vertex_y


def plot_a2_open_interval(records: pd.DataFrame, chinese: FontProperties) -> float:
    frame = records.loc[(records["group"] == "A2") & (records["temperature_c"].isin([300, 325, 350]))].sort_values("temperature_c")
    y325 = float(frame.loc[frame["temperature_c"] == 325, "Y_C4_pct"].iloc[0])
    y350 = float(frame.loc[frame["temperature_c"] == 350, "Y_C4_pct"].iloc[0])
    slope = (y350 - y325) / 25.0
    grid = np.linspace(325, 350, 200)

    figure, axis = plt.subplots(figsize=(8.1, 5.0))
    axis.scatter(frame["temperature_c"], frame["Y_C4_pct"], color="#267D72", s=38, zorder=3, label="A2 实测节点")
    axis.plot(grid, y325 + slope * (grid - 325), color="#267D72", linewidth=2.0, label="325--350 ℃局部线性函数")
    axis.axvline(350, color="#7B2D26", linewidth=1.35, linestyle="--")
    axis.annotate("严格边界 $T=350$ ℃", xy=(350, y350), xytext=(314, 26.15), arrowprops={"arrowstyle": "->", "color": "#7B2D26"}, fontproperties=chinese)
    axis.annotate("上确界逼近方向", xy=(349.4, y350 - 0.2), xytext=(323, 24.0), arrowprops={"arrowstyle": "->", "color": "#267D72"}, fontproperties=chinese)
    label(axis, "温度 $T$ / ℃", r"$\mathrm{C}_4$ 烯烃收率 $Y_{\mathrm{C4}}$ / %", "A2 在严格低温约束下的局部连续化", chinese)
    axis.legend(prop=chinese, frameon=False, loc="upper left")
    figure.tight_layout()
    figure.savefig(FIGURE_DIR / "q3_a2_strict_low_temperature.png", bbox_inches="tight")
    plt.close(figure)
    return slope


def design_matrix(frame: pd.DataFrame, model: str) -> np.ndarray:
    u = (frame["temperature_c"].to_numpy(dtype=float) - 325.0) / 100.0
    columns = [np.ones(len(frame)), u, u**2]
    if model in {"main_effects", "two_interactions"}:
        columns.extend(frame[column].to_numpy(dtype=float) for column in ["co_loading_wt_pct", "total_catalyst_mg", "rho", "ethanol_feed_ml_min", "packing_mode"])
    if model == "two_interactions":
        columns.extend([u * frame["co_loading_wt_pct"].to_numpy(dtype=float), u * frame["ethanol_feed_ml_min"].to_numpy(dtype=float)])
    return np.column_stack(columns)


def loco_metrics(proxy: pd.DataFrame, model: str) -> dict[str, float | int]:
    predictions: list[float] = []
    observed: list[float] = []
    a3_400_prediction: float | None = None
    for held_group in COMBINATIONS:
        test = proxy.loc[proxy["group"] == held_group]
        if test.empty:
            continue
        train = proxy.loc[proxy["group"] != held_group]
        beta = np.linalg.lstsq(design_matrix(train, model), train["Y_C4_pct"].to_numpy(dtype=float), rcond=None)[0]
        predicted = design_matrix(test, model) @ beta
        predictions.extend(predicted.tolist())
        observed.extend(test["Y_C4_pct"].to_numpy(dtype=float).tolist())
        if held_group == "A3":
            target = test.loc[test["temperature_c"] == 400]
            a3_400_prediction = float((design_matrix(target, model) @ beta).item())
    errors = np.asarray(observed) - np.asarray(predictions)
    return {
        "parameter_count": design_matrix(proxy.iloc[:1], model).shape[1],
        "loco_rmse_pct_point": float(math.sqrt(np.mean(errors**2))),
        "loco_mae_pct_point": float(np.mean(np.abs(errors))),
        "a3_400_loco_prediction_pct": float(a3_400_prediction),
    }


def make_tables(records: pd.DataFrame, proxy_metrics: dict[str, dict[str, float | int]], vertex_t: float, vertex_y: float, a2_slope: float) -> dict[str, object]:
    ranking = records.sort_values(["Y_C4_pct", "group", "temperature_c"], ascending=[False, True, True], key=lambda s: group_order(s) if s.name == "group" else s).reset_index(drop=True)
    ranking.insert(0, "rank", np.arange(1, len(ranking) + 1))
    ranking[["rank", "group", "temperature_c", "X_EtOH_display_pct", "S_C4_pct", "Y_C4_pct"]].rename(columns={"X_EtOH_display_pct": "X_EtOH_pct"}).to_csv(TABLE_DIR / "q3_all_observed_yield_ranking.csv", index=False, encoding="utf-8-sig", float_format="%.6f")
    ranking_rows = [[int(row.rank), row.group, f"{row.temperature_c:.0f}", f"{row.X_EtOH_display_pct:.2f}", f"{row.S_C4_pct:.2f}", f"{row.Y_C4_pct:.6f}"] for row in ranking.itertuples(index=False)]
    write_longtable("q3_all_observed_yield_ranking", ["排名", "组合", "$T$/℃", "$X_{\\mathrm{EtOH}}$/\\%", "$S_{\\mathrm{C4}}$/\\%", "$Y_{\\mathrm{C4}}$/\\%"], ranking_rows, "全部实测节点的 $\\mathrm{C}_4$ 烯烃收率排序", "tab:q3-all-ranking", "C{0.07\\textwidth}C{0.08\\textwidth}C{0.11\\textwidth}C{0.18\\textwidth}C{0.18\\textwidth}C{0.18\\textwidth}")
    write_longtable("q3_observed_top5", ["排名", "组合--温度", "$Y_{\\mathrm{C4}}$/\\%"], [[int(row.rank), f"{row.group}--{row.temperature_c:.0f} ℃", f"{row.Y_C4_pct:.2f}"] for row in ranking.head(5).itertuples(index=False)], "全部实测节点中收率最高的 5 个条件", "tab:q3-observed-top5", "C{0.12\\textwidth}C{0.35\\textwidth}C{0.22\\textwidth}")

    strict = records.loc[records["temperature_c"] < 350].sort_values(["Y_C4_pct", "group", "temperature_c"], ascending=[False, True, True], key=lambda s: group_order(s) if s.name == "group" else s).reset_index(drop=True)
    strict.insert(0, "rank", np.arange(1, len(strict) + 1))
    strict[["rank", "group", "temperature_c", "X_EtOH_display_pct", "S_C4_pct", "Y_C4_pct"]].rename(columns={"X_EtOH_display_pct": "X_EtOH_pct"}).to_csv(TABLE_DIR / "q3_strict_lt350_yield_ranking.csv", index=False, encoding="utf-8-sig", float_format="%.6f")
    strict_rows = [[int(row.rank), row.group, f"{row.temperature_c:.0f}", f"{row.X_EtOH_display_pct:.2f}", f"{row.S_C4_pct:.2f}", f"{row.Y_C4_pct:.6f}"] for row in strict.itertuples(index=False)]
    write_longtable("q3_strict_lt350_yield_ranking", ["排名", "组合", "$T$/℃", "$X_{\\mathrm{EtOH}}$/\\%", "$S_{\\mathrm{C4}}$/\\%", "$Y_{\\mathrm{C4}}$/\\%"], strict_rows, "严格 $T<350$ ℃ 的实测节点收率排序", "tab:q3-strict-ranking", "C{0.07\\textwidth}C{0.08\\textwidth}C{0.11\\textwidth}C{0.18\\textwidth}C{0.18\\textwidth}C{0.18\\textwidth}")
    write_longtable("q3_strict_lt350_top5", ["排名", "组合--温度", "$Y_{\\mathrm{C4}}$/\\%"], [[int(row.rank), f"{row.group}--{row.temperature_c:.0f} ℃", f"{row.Y_C4_pct:.2f}"] for row in strict.head(5).itertuples(index=False)], "严格 $T<350$ ℃ 时收率最高的 5 个实测条件", "tab:q3-strict-top5", "C{0.12\\textwidth}C{0.35\\textwidth}C{0.22\\textwidth}")

    common = sorted(set.intersection(*(set(frame["temperature_c"]) for _, frame in records.groupby("group", sort=False))))
    common_rows = []
    for temperature in common:
        row = records.loc[records["temperature_c"] == temperature].sort_values(["Y_C4_pct", "group"], ascending=[False, True], key=lambda s: group_order(s) if s.name == "group" else s).iloc[0]
        common_rows.append({"temperature_c": float(temperature), "group": row["group"], "Y_C4_pct": float(row["Y_C4_pct"])})
    common_frame = pd.DataFrame(common_rows)
    common_frame.to_csv(TABLE_DIR / "q3_common_temperature_best.csv", index=False, encoding="utf-8-sig", float_format="%.6f")
    write_longtable("q3_common_temperature_best", ["共同温度 / ℃", "最高实测组合", "$Y_{\\mathrm{C4}}$/\\%"], [[f"{row.temperature_c:.0f}", row.group, f"{row.Y_C4_pct:.2f}"] for row in common_frame.itertuples(index=False)], "21 组共同实测温度上的最高收率组合", "tab:q3-common-temperature", "C{0.23\\textwidth}C{0.26\\textwidth}C{0.22\\textwidth}")

    metric_names = [("temperature_only", "仅温度 $u+u^2$"), ("main_effects", "低阶主效应代理"), ("two_interactions", "主效应加两个温度交互")]
    proxy_rows = [[name, int(proxy_metrics[key]["parameter_count"]), f"{proxy_metrics[key]['loco_rmse_pct_point']:.4f}", f"{proxy_metrics[key]['loco_mae_pct_point']:.4f}", f"{proxy_metrics[key]['a3_400_loco_prediction_pct']:.2f}"] for key, name in metric_names]
    pd.DataFrame([{"model": key, **metrics} for key, metrics in proxy_metrics.items()]).to_csv(TABLE_DIR / "q3_proxy_loco_qualification.csv", index=False, encoding="utf-8-sig", float_format="%.6f")
    write_longtable("q3_proxy_loco_qualification", ["候选模型", "参数数", "LOCO RMSE", "LOCO MAE", "A3--400 ℃留组预测 / \\%"], proxy_rows, "候选因素级代理的按催化剂组合留出检验", "tab:q3-proxy-loco", "L{0.28\\textwidth}C{0.11\\textwidth}C{0.16\\textwidth}C{0.16\\textwidth}C{0.19\\textwidth}")

    summary = [
        ["无温度限制", "A3--400 ℃", "44.72", "全部实测节点排序与局部连续化支持范围"],
        ["严格 $T<350$ ℃的实测最优", "A2--325 ℃", "17.26", "严格可行的实测节点"],
        ["严格 $T<350$ ℃的局部模型上确界", "$T\\to350^-$（A2）", "26.54", "开区间边界方向，不是可取得最大值"],
    ]
    write_longtable("q3_conclusion_summary", ["情形", "条件", "$Y_{\\mathrm{C4}}$/\\%", "结果性质"], summary, "问题三的收率优化结果汇总", "tab:q3-conclusion-summary", "L{0.27\\textwidth}L{0.24\\textwidth}C{0.15\\textwidth}L{0.23\\textwidth}")
    return {"ranking": ranking, "strict": strict, "common": common_frame, "vertex_t": vertex_t, "vertex_y": vertex_y, "a2_slope": a2_slope}


def validate(records: pd.DataFrame, products: dict[str, object], proxy_metrics: dict[str, dict[str, float | int]]) -> dict[str, object]:
    ranking = products["ranking"]
    strict = products["strict"]
    common = products["common"]
    observed = ranking.iloc[0]
    strict_observed = strict.iloc[0]
    checks: dict[str, bool] = {
        "114_records": len(records) == 114,
        "21_combinations": records["group"].nunique() == 21,
        "all_yields_in_0_100": bool(records["Y_C4_pct"].between(0, 100).all()),
        "yield_no_greater_than_conversion": bool((records["Y_C4_pct"] <= records["X_EtOH_display_pct"] + 1e-10).all()),
        "a11_rho_undefined": bool(records.loc[records["group"] == "A11", "rho"].isna().all()),
        "observed_best": observed["group"] == EXPECTED["observed_best"][0] and observed["temperature_c"] == EXPECTED["observed_best"][1] and abs(observed["Y_C4_pct"] - EXPECTED["observed_best"][2]) < 5e-5,
        "strict_observed_best": strict_observed["group"] == EXPECTED["strict_observed_best"][0] and strict_observed["temperature_c"] == EXPECTED["strict_observed_best"][1] and abs(strict_observed["Y_C4_pct"] - EXPECTED["strict_observed_best"][2]) < 5e-5,
        "common_temperatures": common["temperature_c"].tolist() == [250.0, 275.0, 300.0, 350.0],
        "common_temperature_winners": common["group"].tolist() == ["A7", "A2", "A2", "A2"],
        "a2_boundary_value": abs(float(records.loc[(records["group"] == "A2") & (records["temperature_c"] == 350), "Y_C4_pct"].iloc[0]) - EXPECTED["strict_supremum"]) < 5e-5,
        "a2_slope": abs(float(products["a2_slope"]) - 0.37110096) < 5e-7,
        "quadratic_pressure_vertex": abs(float(products["vertex_t"]) - EXPECTED["quadratic_vertex_temperature"]) < 0.02 and abs(float(products["vertex_y"]) - EXPECTED["quadratic_vertex_yield"]) < 0.02,
    }
    proxy = records.loc[(~records["is_a11_special"]) & (records["temperature_c"] <= 400)].copy()
    checks["proxy_sample_shape"] = len(proxy) == EXPECTED["proxy_rows"] and proxy["group"].nunique() == EXPECTED["proxy_groups"]
    for model, (count, rmse, mae, a3_prediction) in EXPECTED["loco"].items():
        result = proxy_metrics[model]
        checks[f"{model}_metrics"] = int(result["parameter_count"]) == count and abs(float(result["loco_rmse_pct_point"]) - rmse) < 5e-4 and abs(float(result["loco_mae_pct_point"]) - mae) < 5e-4 and abs(float(result["a3_400_loco_prediction_pct"]) - a3_prediction) < 0.01
    return {
        "status": "PASS" if all(checks.values()) else "DISCREPANCY",
        "checks": checks,
        "source_hashes": {str(path.relative_to(ROOT)): hashlib.sha256(path.read_bytes()).hexdigest() for path in [RAW_PATH, HANDOFF_PATH] if path.exists()},
        "proxy_metrics": proxy_metrics,
    }


def write_log(report: dict[str, object]) -> None:
    (LOG_DIR / "q3_validation.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=lambda value: value.item() if isinstance(value, np.generic) else str(value)) + "\n", encoding="utf-8")
    lines = ["# 问题三结果核验", "", f"状态：**{report['status']}**", ""]
    lines.extend(f"- {'通过' if passed else '失败'}：`{name}`" for name, passed in report["checks"].items())
    lines += ["", r"说明：收率按 $Y_{\mathrm{C4}}=X_{\mathrm{EtOH}}S_{\mathrm{C4}}/100$ 计算；未实测温度没有补值。候选因素级代理仅用于检验其是否具备为未实验配方排序的能力。"]
    (LOG_DIR / "q3_validation.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--strict", action="store_true", help="核验失败时返回非零状态。")
    args = parser.parse_args()
    for directory in [TABLE_DIR, FIGURE_DIR, LOG_DIR]:
        directory.mkdir(parents=True, exist_ok=True)
    records, _ = load_data(RAW_PATH)
    display_x = displayed_conversion_values(RAW_PATH)
    records["X_EtOH_display_pct"] = [display_x[(group, float(temperature))] for group, temperature in zip(records["group"], records["temperature_c"])]
    records["Y_C4_pct"] = records["X_EtOH_display_pct"] * records["S_C4_pct"] / 100.0
    records.to_csv(OUTPUT_DIR / "q3_observed_records.csv", index=False, encoding="utf-8-sig", float_format="%.8f")
    times, chinese = style_plots()
    _ = times
    plot_all_trajectories(records, chinese)
    vertex_t, vertex_y = plot_a3_sensitivity(records, chinese)
    a2_slope = plot_a2_open_interval(records, chinese)
    proxy = records.loc[(~records["is_a11_special"]) & (records["temperature_c"] <= 400)].copy()
    proxy_metrics = {model: loco_metrics(proxy, model) for model in EXPECTED["loco"]}
    products = make_tables(records, proxy_metrics, vertex_t, vertex_y, a2_slope)
    report = validate(records, products, proxy_metrics)
    write_log(report)
    print(json.dumps({"status": report["status"], "checks": report["checks"], "proxy_metrics": proxy_metrics}, ensure_ascii=False, indent=2, default=lambda value: value.item() if isinstance(value, np.generic) else str(value)))
    return 0 if report["status"] == "PASS" or not args.strict else 1


if __name__ == "__main__":
    raise SystemExit(main())
